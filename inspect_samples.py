#!/usr/bin/env python3
"""
Comprehensive inspection of the three sample Excel files.
Uses openpyxl for detailed structural analysis.
"""

import openpyxl
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import json

def inspect_workbook(file_path):
    """Inspect an Excel workbook in detail."""
    print(f"\n{'='*100}")
    print(f"INSPECTING: {file_path}")
    print(f"{'='*100}\n")
    
    try:
        wb = load_workbook(file_path)
        
        # 1. Workbook sheet names
        print(f"1. SHEET NAMES: {wb.sheetnames}")
        print(f"   Active Sheet: {wb.active.title}\n")
        
        # Process each sheet
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            print(f"\n--- WORKSHEET: '{sheet_name}' ---")
            print(f"Dimensions: {ws.dimensions}")
            print(f"Max Row: {ws.max_row}, Max Column: {ws.max_column}\n")
            
            # 3. Extract headers (first row)
            print("2. HEADERS (ROW 1):")
            headers = []
            for col_idx, cell in enumerate(ws[1], 1):
                header_value = cell.value
                headers.append(header_value)
                print(f"   Column {col_idx} ({get_column_letter(col_idx)}): '{header_value}' (type: {type(header_value).__name__})")
            
            print("\n3. FIRST 10 DATA ROWS:")
            print("-" * 100)
            
            # 4 & 5. First 10 rows with cell data types
            for row_idx in range(2, min(12, ws.max_row + 1)):
                print(f"\nRow {row_idx}:")
                for col_idx, cell in enumerate(ws[row_idx], 1):
                    cell_value = cell.value
                    cell_type = type(cell_value).__name__
                    data_type = cell.data_type
                    number_format = cell.number_format
                    
                    # 7. Check for hyperlinks
                    hyperlink_info = ""
                    if cell.hyperlink:
                        hyperlink_info = f" [HYPERLINK: {cell.hyperlink.target}]"
                    
                    # 8. Check for blank/null
                    if cell_value is None:
                        value_str = "[BLANK]"
                    else:
                        value_str = str(cell_value)
                    
                    print(f"  {get_column_letter(col_idx)}{row_idx}: {value_str}")
                    print(f"         Type: {cell_type}, DataType: {data_type}, Format: {number_format}{hyperlink_info}")
            
            # 9. Check for merged cells
            print(f"\n4. MERGED CELLS:")
            if ws.merged_cells:
                for merged_range in ws.merged_cells.ranges:
                    print(f"   {merged_range}")
            else:
                print("   None")
            
            # 11. Check for duplicate column names
            print(f"\n5. DUPLICATE COLUMN CHECK:")
            from collections import Counter
            header_counts = Counter(headers)
            duplicates = {h: count for h, count in header_counts.items() if count > 1 and h is not None}
            if duplicates:
                print(f"   Found duplicates: {duplicates}")
                for header, count in duplicates.items():
                    positions = [i+1 for i, h in enumerate(headers) if h == header]
                    print(f"   '{header}' appears at columns: {positions}")
            else:
                print("   No duplicate column names found")
            
            # 6. Date values and formats (look for date patterns)
            print(f"\n6. DATE AND NUMBER FORMATS:")
            date_cells = []
            for row in ws.iter_rows(min_row=2, max_row=min(12, ws.max_row)):
                for cell in row:
                    if cell.data_type == 'd' or 'date' in (cell.number_format or '').lower() or 'yyyy' in (cell.number_format or '').lower():
                        date_cells.append((cell.coordinate, cell.value, cell.number_format))
                    elif isinstance(cell.value, (int, float)) and 'percent' in (cell.number_format or '').lower():
                        print(f"   Percentage: {cell.coordinate} = {cell.value} (format: {cell.number_format})")
            
            if date_cells:
                for coord, val, fmt in date_cells:
                    print(f"   Date Cell {coord}: {val} (format: {fmt})")
            
            print("\n" + "="*100)
    
    except Exception as e:
        print(f"ERROR inspecting {file_path}: {e}")
        import traceback
        traceback.print_exc()

def main():
    files = [
        "crowdlog_sample.xlsx",
        "client_sample.xlsx",
        "monthly_report_example.xlsx"
    ]
    
    for file_path in files:
        inspect_workbook(file_path)
    
    print("\n\n" + "="*100)
    print("INSPECTION COMPLETE")
    print("="*100)

if __name__ == "__main__":
    main()
