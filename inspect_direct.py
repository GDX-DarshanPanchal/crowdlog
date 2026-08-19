#!/usr/bin/env python3
"""
Direct inspection of Excel files using openpyxl.
Outputs results to console and file.
"""

import openpyxl
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from datetime import datetime

def inspect_workbook(file_path):
    """Inspect an Excel workbook in detail."""
    output_lines = []
    
    output_lines.append(f"\n{'='*120}")
    output_lines.append(f"FILE: {file_path}")
    output_lines.append(f"{'='*120}\n")
    
    try:
        wb = load_workbook(file_path, data_only=False)
        
        # Sheet names
        output_lines.append(f"SHEET NAMES: {wb.sheetnames}")
        output_lines.append(f"Active Sheet: {wb.active.title}\n")
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            output_lines.append(f"\n{'─'*120}")
            output_lines.append(f"WORKSHEET: '{sheet_name}'")
            output_lines.append(f"Dimensions: {ws.dimensions}")
            output_lines.append(f"Max Row: {ws.max_row}, Max Column: {ws.max_column}")
            output_lines.append(f"{'─'*120}\n")
            
            # Headers
            output_lines.append("COLUMN HEADERS (Row 1):")
            headers = []
            for col_idx, cell in enumerate(ws[1], 1):
                header_value = cell.value
                headers.append(header_value)
                output_lines.append(f"  Col {col_idx:2d} ({get_column_letter(col_idx):2s}): {repr(header_value)}")
            
            # Check for duplicate headers
            from collections import Counter
            header_counts = Counter(h for h in headers if h is not None)
            duplicates = {h: count for h, count in header_counts.items() if count > 1}
            if duplicates:
                output_lines.append(f"\n⚠️  DUPLICATE COLUMN NAMES FOUND:")
                for header, count in duplicates.items():
                    positions = [i+1 for i, h in enumerate(headers) if h == header]
                    output_lines.append(f"    '{header}' appears {count} times at columns: {positions}")
            
            # Sample data rows
            output_lines.append(f"\n\nFIRST 5 DATA ROWS (Rows 2-6):")
            output_lines.append("─" * 120)
            
            for row_idx in range(2, min(7, ws.max_row + 1)):
                output_lines.append(f"\nRow {row_idx}:")
                for col_idx, cell in enumerate(ws[row_idx], 1):
                    cell_value = cell.value
                    cell_type = type(cell_value).__name__
                    data_type = cell.data_type
                    number_format = cell.number_format
                    
                    # Hyperlink info
                    hyperlink_info = ""
                    if cell.hyperlink:
                        hyperlink_info = f" → HYPERLINK: {cell.hyperlink.target}"
                    
                    # Format value for display
                    if cell_value is None:
                        value_str = "[EMPTY]"
                    elif isinstance(cell_value, datetime):
                        value_str = f"{cell_value.strftime('%Y-%m-%d %H:%M:%S')}"
                    else:
                        value_str = str(cell_value)[:80]
                    
                    col_letter = get_column_letter(col_idx)
                    header = headers[col_idx - 1] if col_idx <= len(headers) else "?"
                    
                    output_lines.append(f"  {col_letter}{row_idx} [{header:20s}]: {value_str}")
                    output_lines.append(f"        (Type: {cell_type}, DataType: {data_type}, Format: {number_format}){hyperlink_info}")
            
            # Merged cells
            output_lines.append(f"\n\nMERGED CELLS:")
            if ws.merged_cells:
                for merged_range in ws.merged_cells.ranges:
                    output_lines.append(f"  {merged_range}")
            else:
                output_lines.append("  None")
            
            # Check for hyperlinks in all data rows
            output_lines.append(f"\n\nHYPERLINKS FOUND:")
            hyperlinks_found = []
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                for cell in row:
                    if cell.hyperlink:
                        header = headers[cell.column - 1] if cell.column <= len(headers) else "?"
                        hyperlinks_found.append((cell.coordinate, header, cell.value, cell.hyperlink.target))
            
            if hyperlinks_found:
                for coord, header, value, target in hyperlinks_found[:10]:  # Show first 10
                    output_lines.append(f"  {coord} [{header}]: {value} → {target}")
                if len(hyperlinks_found) > 10:
                    output_lines.append(f"  ... and {len(hyperlinks_found) - 10} more hyperlinks")
            else:
                output_lines.append("  None")
            
            # Date and number formats
            output_lines.append(f"\n\nDATE & NUMBER FORMATS:")
            date_formats = set()
            for row in ws.iter_rows(min_row=2, max_row=min(20, ws.max_row)):
                for cell in row:
                    fmt = cell.number_format or ""
                    if 'date' in fmt.lower() or 'yyyy' in fmt.lower() or cell.data_type == 'd':
                        date_formats.add(fmt)
            
            if date_formats:
                for fmt in sorted(date_formats):
                    output_lines.append(f"  {fmt}")
            else:
                output_lines.append("  No date formats detected")
            
            output_lines.append(f"\n")
    
    except Exception as e:
        output_lines.append(f"ERROR: {e}")
        import traceback
        output_lines.append(traceback.format_exc())
    
    return "\n".join(output_lines)

def main():
    files = [
        "crowdlog_sample.xlsx",
        "client_sample.xlsx",
        "monthly_report_example.xlsx"
    ]
    
    all_output = []
    all_output.append("EXCEL FILE INSPECTION REPORT")
    all_output.append(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    all_output.append("=" * 120)
    
    for file_path in files:
        output = inspect_workbook(file_path)
        all_output.append(output)
    
    all_output.append("\n\n" + "=" * 120)
    all_output.append("INSPECTION COMPLETE")
    all_output.append("=" * 120)
    
    full_output = "\n".join(all_output)
    
    # Print to console
    print(full_output)
    
    # Save to file
    with open("INSPECTION_RESULTS.txt", "w") as f:
        f.write(full_output)
    
    print("\n✓ Results saved to INSPECTION_RESULTS.txt")

if __name__ == "__main__":
    main()
