from __future__ import annotations

import csv
from pathlib import Path
from typing import Any, Iterator


def iter_records(path: Path) -> Iterator[tuple[int, dict[str, Any], dict[str, str]]]:
    """Yield row number, values, and hyperlink targets without changing input."""
    if path.suffix.lower() == ".csv":
        with path.open(encoding="utf-8-sig", newline="") as handle:
            for number, row in enumerate(csv.DictReader(handle), 2):
                yield number, dict(row), {}
        return
    from openpyxl import load_workbook
    workbook = load_workbook(path, data_only=True, read_only=False)
    try:
        sheet = workbook.active
        headers = [str(c.value).strip() if c.value is not None else "" for c in sheet[1]]
        for number, cells in enumerate(sheet.iter_rows(min_row=2), 2):
            values = {headers[i]: cell.value for i, cell in enumerate(cells) if i < len(headers) and headers[i]}
            links = {headers[i]: cell.hyperlink.target for i, cell in enumerate(cells) if i < len(headers) and headers[i] and cell.hyperlink}
            if any(value not in (None, "") for value in values.values()):
                yield number, values, links
    finally:
        workbook.close()
