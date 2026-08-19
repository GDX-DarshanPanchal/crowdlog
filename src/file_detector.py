from __future__ import annotations

import csv
from pathlib import Path


CROWDLOG_HEADERS = {"timesheet_date", "member_name", "minutes", "memo"}
CLIENT_HEADERS = {"Task", "Ticket", "Task Title", "Ticket Logged Date", "Start Date", "End Date"}


def headers_for(path: Path) -> set[str]:
    if path.suffix.lower() == ".csv":
        with path.open(encoding="utf-8-sig", newline="") as handle:
            return {str(x).strip() for x in next(csv.reader(handle), [])}
    if path.suffix.lower() == ".xlsx":
        from openpyxl import load_workbook
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            return {str(cell.value).strip() for cell in next(workbook.active.iter_rows()) if cell.value is not None}
        finally:
            workbook.close()
    return set()


def detect_files(folder: Path) -> tuple[Path, Path]:
    crowdlog: list[Path] = []
    client: list[Path] = []
    for path in sorted(folder.iterdir()) if folder.exists() else []:
        if path.suffix.lower() not in {".xlsx", ".csv"} or path.name.startswith("~$"):
            continue
        headers = headers_for(path)
        if CROWDLOG_HEADERS <= headers:
            crowdlog.append(path)
        if CLIENT_HEADERS <= headers:
            client.append(path)
    if len(crowdlog) != 1:
        raise ValueError("Could not identify one Crowdlog file. Please make sure it contains: timesheet_date, member_name, minutes and memo.")
    if len(client) != 1:
        raise ValueError("Could not identify one Client/JIRA file. Please make sure it contains: Task, Ticket, Task Title, Ticket Logged Date, Start Date and End Date.")
    return crowdlog[0], client[0]
