from __future__ import annotations

from pathlib import Path

from .action_summary import action_summary
from .aggregator import hours
from .categorizer import suggestions
from .date_utils import week_in_month
from .models import ReportRow, ReviewItem

HEADERS = ["Resource", "Week no_", "Date", "Resource", "Ticket", "Issue", "Action to Check", "Resolution",
           "Log", "Issue Log Date", "Start Date", "End Date/ Resolved date", "Status", "Final log", "Task type",
           "Billable status", "Suggested Task type", "Suggested Billable status"]
REVIEW_HEADERS = ["Reason", "Source", "Source Row", "Employee", "Date", "Ticket", "Details", "Original Data"]


def unique_output(folder: Path, month: str) -> Path:
    folder.mkdir(parents=True, exist_ok=True)
    base = folder / f"{month}_monthly_report.xlsx"
    if not base.exists():
        return base
    index = 2
    while (candidate := folder / f"{month}_monthly_report_{index}.xlsx").exists():
        index += 1
    return candidate


def write_report(path: Path, rows: list[ReportRow], reviews: list[ReviewItem], settings: dict) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    workbook = Workbook()
    report = workbook.active
    report.title = "Monthly Report"
    report.append(HEADERS)
    for item in rows:
        suggested_task, suggested_billable = suggestions(item.billable_status, item.client.task, settings)
        value_hours = hours(item.minutes)
        report.append([settings["resource_group"], week_in_month(item.work_date), item.work_date,
                       settings["resource_aliases"].get(item.employee, item.employee), item.ticket, item.client.title,
                       action_summary(item.dates, item.memos, item.ticket), item.ticket, value_hours,
                       item.client.logged_date, item.client.start_date, item.client.end_date, item.status, value_hours,
                       item.task_type, item.billable_status, suggested_task, suggested_billable])
    _format_sheet(report, HEADERS, Font, PatternFill, Alignment, get_column_letter)
    for row in report.iter_rows(min_row=2):
        for col in (3, 10, 11, 12):
            row[col - 1].number_format = "DD-MMM-YYYY"
        for col in (9, 14):
            row[col - 1].number_format = "0.####"
        row[6].alignment = Alignment(wrap_text=True, vertical="top")

    review = workbook.create_sheet("Review Needed")
    review.append(REVIEW_HEADERS)
    for item in reviews:
        review.append([item.reason, item.source, item.row, item.employee, item.date, item.ticket,
                       item.details, item.original_data])
    _format_sheet(review, REVIEW_HEADERS, Font, PatternFill, Alignment, get_column_letter)
    review.column_dimensions["H"].width = 70
    for row in review.iter_rows(min_row=2):
        row[7].alignment = Alignment(wrap_text=True, vertical="top")
    workbook.save(path)


def _format_sheet(sheet, headers, Font, PatternFill, Alignment, get_column_letter) -> None:
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{max(sheet.max_row, 1)}"
    fill = PatternFill("solid", fgColor="1F4E78")
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = fill
        cell.alignment = Alignment(vertical="center")
    widths = {1: 13, 2: 10, 3: 16, 4: 20, 5: 14, 6: 28, 7: 55, 8: 14, 9: 10,
              10: 18, 11: 16, 12: 24, 13: 15, 14: 12, 15: 25, 16: 27, 17: 25, 18: 27}
    for index in range(1, len(headers) + 1):
        sheet.column_dimensions[get_column_letter(index)].width = widths.get(index, 18)
