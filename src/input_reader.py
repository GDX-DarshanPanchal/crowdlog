"""Read and validate input XLSX files."""

import os
from openpyxl import load_workbook
from typing import List, Dict, Optional
from src.exceptions import (
    FileNotFoundError as CrowdlogFileNotFoundError,
    InvalidXLSXError,
    MissingColumnError,
    InputValidationError,
)
from src.config import REQUIRED_CROWDLOG_COLUMNS, REQUIRED_CLIENT_COLUMNS


class InputReader:
    """Read and validate Crowdlog and Client XLSX files."""
    
    def __init__(self):
        """Initialize input reader."""
        pass
    
    def read_crowdlog(self, file_path: str) -> List[Dict]:
        """Read Crowdlog XLSX file.
        
        Args:
            file_path: Path to Crowdlog XLSX file
            
        Returns:
            List of Crowdlog records as dictionaries
            
        Raises:
            FileNotFoundError: If file does not exist
            InvalidXLSXError: If file is not valid XLSX
            MissingColumnError: If required columns are missing
            InputValidationError: If validation fails
        """
        # Validate file exists
        if not os.path.exists(file_path):
            raise CrowdlogFileNotFoundError(f"Crowdlog file not found: {file_path}")
        
        # Read workbook
        try:
            wb = load_workbook(file_path, data_only=True)
        except Exception as e:
            raise InvalidXLSXError(f"Cannot read Crowdlog file as XLSX: {e}")
        
        # Get active sheet
        ws = wb.active
        if not ws:
            raise InputValidationError("Crowdlog file has no active worksheet")
        
        # Extract headers
        headers = []
        for cell in ws[1]:
            headers.append(cell.value)
        
        if not headers:
            raise InputValidationError("Crowdlog file has no headers")
        
        # Validate required columns
        for required_col in REQUIRED_CROWDLOG_COLUMNS:
            if required_col not in headers:
                raise MissingColumnError(
                    f"Crowdlog file missing required column: '{required_col}'. "
                    f"Found columns: {headers}"
                )
        
        # Read data rows
        records = []
        for row_idx in range(2, ws.max_row + 1):
            row_data = {}
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                row_data[header] = cell.value
            
            # Skip empty rows
            if any(row_data.values()):
                records.append(row_data)
        
        if not records:
            raise InputValidationError("Crowdlog file has no data rows")
        
        return records
    
    def read_client(self, file_path: str) -> List[Dict]:
        """Read Client XLSX file.
        
        Args:
            file_path: Path to Client XLSX file
            
        Returns:
            List of Client records as dictionaries
            
        Raises:
            FileNotFoundError: If file does not exist
            InvalidXLSXError: If file is not valid XLSX
            MissingColumnError: If required columns are missing
            InputValidationError: If validation fails
        """
        # Validate file exists
        if not os.path.exists(file_path):
            raise CrowdlogFileNotFoundError(f"Client file not found: {file_path}")
        
        # Read workbook
        try:
            wb = load_workbook(file_path, data_only=True)
        except Exception as e:
            raise InvalidXLSXError(f"Cannot read Client file as XLSX: {e}")
        
        # Get active sheet
        ws = wb.active
        if not ws:
            raise InputValidationError("Client file has no active worksheet")
        
        # Extract headers
        headers = []
        for cell in ws[1]:
            headers.append(cell.value)
        
        if not headers:
            raise InputValidationError("Client file has no headers")
        
        # Validate required columns
        for required_col in REQUIRED_CLIENT_COLUMNS:
            if required_col not in headers:
                raise MissingColumnError(
                    f"Client file missing required column: '{required_col}'. "
                    f"Found columns: {headers}"
                )
        
        # Read data rows and extract hyperlinks
        records = []
        for row_idx in range(2, ws.max_row + 1):
            row_data = {}
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                value = cell.value
                
                # Check for hyperlink in Ticket column
                if header == 'Ticket' and cell.hyperlink:
                    # Store both the display value and the hyperlink target
                    row_data[f'{header}_hyperlink'] = cell.hyperlink.target
                    row_data[header] = value
                else:
                    row_data[header] = value
            
            # Skip empty rows
            if any(row_data.get(k) for k in REQUIRED_CLIENT_COLUMNS):
                records.append(row_data)
        
        if not records:
            raise InputValidationError("Client file has no data rows")
        
        return records
    
    def validate_crowdlog_record(self, record: Dict) -> bool:
        """Validate a single Crowdlog record.
        
        Args:
            record: Crowdlog record dictionary
            
        Returns:
            True if record is valid
        """
        # Check required fields
        for col in REQUIRED_CROWDLOG_COLUMNS:
            if col not in record or not record[col]:
                return False
        
        return True
    
    def validate_client_record(self, record: Dict) -> bool:
        """Validate a single Client record.
        
        Args:
            record: Client record dictionary
            
        Returns:
            True if record is valid
        """
        # Check required fields
        for col in REQUIRED_CLIENT_COLUMNS:
            if col not in record or not record[col]:
                return False
        
        return True
