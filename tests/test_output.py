from datetime import date
from decimal import Decimal

from openpyxl import load_workbook

from src.models import ClientTicket, ReportRow
from src.report_generator import HEADERS, write_report


def test_output_contract(tmp_path):
    settings = {"resource_group": "GDX", "resource_aliases": {"Panchal Darshan": "Darshan"},
                "task_type_aliases": {"EC Operation Support": "Operations"},
                "billable_rules": {"Operations": "Billable"}, "allowed_statuses": ["In Progress"],
                "default_status": "In Progress"}
    client = ClientTicket("OEB-1318", "EC Operation Support", "AW26 Build", date(2026, 6, 19),
                          date(2026, 6, 19), None, 2)
    row = ReportRow("2026-07", "Panchal Darshan", "OEB-1318", date(2026, 7, 13), [date(2026, 7, 13)],
                    ["OEB-1318 review"], Decimal(1110), "Operations", "Billable", client, "In Progress")
    output = tmp_path / "report.xlsx"
    write_report(output, [row], [], settings)
    workbook = load_workbook(output, data_only=True)
    sheet = workbook["Monthly Report"]
    assert workbook.sheetnames == ["Monthly Report", "Review Needed"]
    assert [c.value for c in sheet[1]] == HEADERS
    assert len(HEADERS) == 18 and HEADERS.count("Resource") == 2
    assert sheet.cell(2, 8).value == "OEB-1318"
    assert sheet.cell(2, 9).value == sheet.cell(2, 14).value == 18.5
    assert sheet.cell(2, 15).value == "Operations"
    assert sheet.cell(2, 17).value == "Operations"
    assert isinstance(sheet.cell(2, 3).value, (date, __import__("datetime").datetime))
