"""
Script to inspect the structure of the sample Excel files.
This helps understand the exact column names, data types, and structure.
"""

import openpyxl
import json
from openpyxl import load_workbook

def inspect_excel_file(file_path):
    """
    Inspect an Excel file and extract metadata and sample data.
    """
    print(f"\n{'='*80}")
    print(f"Inspecting: {file_path}")
    print(f"{'='*80}\n")
    
    wb = load_workbook(file_path)
    
    # List all sheets
    print(f"Sheet Names: {wb.sheetnames}\n")
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"\n--- Sheet: {sheet_name} ---")
        print(f"Dimensions: {ws.dimensions}")
        print(f"Max Row: {ws.max_row}, Max Column: {ws.max_column}\n")
        
        # Extract headers and sample data
        print("Headers and Sample Data:")
        print("-" * 80)
        
        for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=min(10, ws.max_row), values_only=False), 1):
            row_data = []
            for cell in row:
                cell_value = cell.value
                cell_type = type(cell_value).__name__ if cell_value is not None else "None"
                row_data.append({
                    "value": str(cell_value),
                    "type": cell_type,
                    "coordinate": cell.coordinate,
                    "data_type": cell.data_type,
                    "number_format": cell.number_format
                })
            
            print(f"\nRow {row_idx}:")
            for i, cell_info in enumerate(row_data):
                if cell_info["value"] and cell_info["value"] != "None":
                    print(f"  Col {i+1} ({cell_info['coordinate']}): {cell_info['value']} ({cell_info['type']}, format: {cell_info['number_format']})")
        
        print("\n" + "-" * 80)

if __name__ == "__main__":
    files_to_inspect = [
        "crowdlog_sample.xlsx",
        "client_sample.xlsx",
        "monthly_report_example.xlsx"
    ]
    
    for file_path in files_to_inspect:
        try:
            inspect_excel_file(file_path)
        except Exception as e:
            print(f"Error inspecting {file_path}: {e}")
